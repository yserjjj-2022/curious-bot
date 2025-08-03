# openalex_explorer.py
import requests
import argparse
import json
import openpyxl

# --- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ---

def save_to_excel_multisheet(filename, sheets_data):
    """Сохраняет данные на разные листы одного Excel-файла."""
    try:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
        for sheet_name, sheet_content in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.append(sheet_content['headers'])
            for row_data in sheet_content['data']: ws.append(row_data)
            for col in ws.columns:
                max_length = 0; column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = min((max_length + 2), 70)
        wb.save(filename)
        print(f"\n✅ Данные успешно сохранены в файл: {filename}")
    except Exception as e: print(f"❌ Ошибка при сохранении в Excel: {e}")

def save_to_excel_single_sheet(filename, headers, data):
    """Сохраняет данные в файл Excel на один лист."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "OpenAlex Search Results"
        ws.append(headers)
        for row_data in data: ws.append(row_data)
        for col in ws.columns:
            max_length = 0; column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min((max_length + 2), 70)
        wb.save(filename)
        print(f"\n✅ Результаты поиска сохранены в файл: {filename}")
    except Exception as e: print(f"❌ Ошибка при сохранении в Excel: {e}")

def fetch_all_entities(entity_type):
    """Загружает все сущности указанного типа (используется для dump)."""
    results = []; url = f"https://api.openalex.org/{entity_type}"; params = {'per-page': 200, 'page': 1}
    print(f"Загружаю все записи для '{entity_type}'... (это может занять время)")
    while True:
        try:
            response = requests.get(url, params=params)
            response.raise_for_status(); data = response.json()
            results.extend(data.get('results', []))
            if not data.get('meta', {}).get('next_page'): break
            params['page'] += 1
        except Exception as e:
            print(f"Ошибка API при загрузке '{entity_type}': {e}"); break
    print(f"-> Загружено {len(results)} записей для '{entity_type}'.")
    return results

# --- ОСНОВНЫЕ ДЕЙСТВИЯ ---

def perform_search(entity_type, search_term, output_file=None):
    """Выполняет поиск по ключевым словам."""
    print(f"\n--- Ищу '{search_term}' в '{entity_type}' ---")
    if entity_type not in ['concepts', 'topics']: print("Ошибка: Поиск поддерживается только для 'concepts' и 'topics'."); return
    url = f"https://api.openalex.org/{entity_type}"; params = {'search': search_term, 'per-page': 200}
    try:
        response = requests.get(url, params=params)
        response.raise_for_status(); results = response.json().get('results', [])
        if not results: print("По вашему запросу ничего не найдено."); return
        data_for_excel = []
        if entity_type == 'topics':
            headers = ["Название", "ID", "Область", "Раздел", "Подраздел"]
            for item in results: data_for_excel.append([item.get('display_name'), item.get('id'), item.get('domain', {}).get('display_name'), item.get('field', {}).get('display_name'), item.get('subfield', {}).get('display_name')])
        else: # concepts
            headers = ["Название", "ID"]
            for item in results: data_for_excel.append([item.get('display_name'), item.get('id')])
        for row in data_for_excel:
            print("-" * 20)
            for header, value in zip(headers, row):
                if value: print(f"{header}: {value}")
        if output_file: save_to_excel_single_sheet(output_file, headers, data_for_excel)
    except Exception as e: print(f"❌ Ошибка во время поиска: {e}")

def perform_dump(output_filename):
    """Выполняет полную выгрузку иерархии в Excel."""
    print("=== Начало полной выгрузки классификаций OpenAlex ==="); all_domains = fetch_all_entities('domains'); all_fields = fetch_all_entities('fields'); all_subfields = fetch_all_entities('subfields'); all_topics = fetch_all_entities('topics'); all_concepts = fetch_all_entities('concepts')
    print("\nСоздаю карты для построения иерархии..."); domains_map = {d['id']: d['display_name'] for d in all_domains}; fields_map = {f['id']: {'name': f['display_name'], 'domain_id': f.get('domain', {}).get('id')} for f in all_fields}; subfields_map = {s['id']: {'name': s['display_name'], 'field_id': s.get('field', {}).get('id')} for s in all_subfields}
    print("Собираю иерархическую таблицу тем..."); topics_data = []
    for topic in all_topics:
        subfield_id = topic.get('subfield', {}).get('id'); subfield_info = subfields_map.get(subfield_id, {}); subfield_name = subfield_info.get('name'); field_id = subfield_info.get('field_id'); field_info = fields_map.get(field_id, {}); field_name = field_info.get('name'); domain_id = field_info.get('domain_id'); domain_name = domains_map.get(domain_id)
        topics_data.append([domain_name, field_name, subfield_name, topic['display_name'], topic['id']])
    concepts_data = [[c['display_name'], c['id']] for c in all_concepts]
    sheets_to_save = {'Topics_Hierarchy': {'headers': ["Область (Domain)", "Раздел (Field)", "Подраздел (Subfield)", "Тема (Topic)", "ID Темы"], 'data': sorted(topics_data, key=lambda x: (x[0] or "", x[1] or "", x[2] or "", x[3] or ""))}, 'Concepts': {'headers': ["Название концепции", "ID Концепции"], 'data': sorted(concepts_data, key=lambda x: x[0])}}
    save_to_excel_multisheet(output_filename, sheets_to_save)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Универсальный исследователь классификаций OpenAlex.")
    parser.add_argument("action", choices=['search', 'dump'], help="Действие: 'search' для поиска или 'dump' для полной выгрузки.")
    parser.add_argument("--entity_type", choices=['concepts', 'topics'], help="Тип сущности для поиска ('search').")
    parser.add_argument("--term", type=str, help="Ключевое слово для поиска ('search').")
    parser.add_argument("--output", type=str, help="Имя файла Excel для сохранения результатов.")
    args = parser.parse_args()
    if args.action == 'search':
        if not (args.entity_type and args.term): parser.error("--entity_type и --term обязательны для действия 'search'.")
        perform_search(args.entity_type, args.term, args.output)
    elif args.action == 'dump':
        if not args.output: parser.error("--output обязателен для действия 'dump'.")
        perform_dump(args.output)
